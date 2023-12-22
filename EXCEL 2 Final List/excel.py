from openpyxl import Workbook, load_workbook

fileToReadEnteriesFrom = "DATA.xlsx"
fileToWriteData = "data_.xlsx"

excel_file = "DATA.xlsx"
workbook = load_workbook(excel_file)
worksheet = workbook.active

def readCell(fileName, cellvalue):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    value = worksheet[cellvalue].value
##    print(value)
    return value

def read_A_Col_data(fileName):
    my_temp_list = []
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    first_column = worksheet['B']
    for x in range(len(first_column)): 
##        print(first_column[x].value)
        my_temp_list.append(first_column[x].value)

    return my_temp_list

#print(read_A_Col_data(fileToReadEnteriesFrom))


def readMaxCol(fileName):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    maxNumberOfColumns = worksheet.max_column
    return maxNumberOfColumns

def readMaxRow(fileName):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    maxNumberOfRow = worksheet.max_row
    return maxNumberOfRow
    
def writeToCellOfEXCEL(fileName, row, col, Value):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    cell = str(col)+str(row)
    worksheet[cell] = Value
    workbook.save(excel_file)
    return True

def writeToNextRow(fileName, col, Value):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    MAXROW = worksheet.max_row
    cell = str(col)+str(MAXROW+1)
    worksheet[cell] = Value
    workbook.save(excel_file)
    return True

def insertListToROW(list1=[], list2=[], fileName=fileToWriteData):
    excel_file = fileName
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
    MAXROW = worksheet.max_row
    finalList = list1 + list2
    for x in range(len(finalList)):
        cell = chr(66+x)+str(MAXROW+1)
        worksheet[cell] = finalList[x]
    workbook.save(fileName)
    return 1
