##import pandas as pd
##
### Specify the file path
##excel_file = 'book1.xlsx'
##
### Read the Excel file into a DataFrame
##df = pd.read_excel(excel_file)
##
##print (df.cell(row = 0, column = 4). value)
##
### Now you can work with the data in the DataFrame (df)


##import openpyxl
##
##wb = openpyxl.load_workbook('book1.xlsx')
##
##ws = wb.get_sheet_by_name('Sheet1')
##
##r = 1
##c = 4
##
####print (ws.cell(row = r, column = c). value)
##print (ws.cell(row = r, column = c). hyperlink.display)
####print (ws.cell(row = r, column = c). hyperlink_rel_id)



from openpyxl import Workbook, load_workbook

# Specify the Excel file you want to write to
excel_file = "book1.xlsx"

# Create a new workbook or load an existing one
try:
    # Load an existing workbook
    workbook = load_workbook(excel_file)
    # Select the worksheet you want to write to (or create a new one)
    worksheet = workbook.active
except FileNotFoundError:
    # Create a new workbook if the file doesn't exist
    workbook = Workbook()
    worksheet = workbook.active

# Write data to the worksheet
for x in range(1, worksheet.max_row+1):
    hyper = worksheet.cell(row = x, column = 4). hyperlink.display
    cell = "F"+str(x)
    worksheet[cell] = hyper

# Save the workbook to the Excel file
workbook.save(excel_file)
